#!/usr/bin/env python3
"""Prepare manifest + in-repo MP3 assets for Nonword Repetition scorer."""
from __future__ import annotations

import json
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parent
DATA_DIR = APP_ROOT / "data"
AUDIO_DIR = DATA_DIR / "audio"

IMMEDIATE_SRC = PROJECT_ROOT / "Analysis" / "BehavioralData" / "ImmediateData" / "NonwordRepetition"
DELAYED_SRC = PROJECT_ROOT / "Analysis" / "BehavioralData" / "DelayedData" / "NonwordRepetition"
SCORES_XLSX = PROJECT_ROOT / "Analysis" / "BehavioralData" / "NonwordRepetition.xlsx"
MANIFEST_PATH = DATA_DIR / "participants.json"

DIR_PATTERN = re.compile(r"^(\d+)_")
SLIDE_PATTERN = re.compile(r"_slide(\d+)\.wav$", re.IGNORECASE)

SCORED_SLIDES = [1, 2, 4, 6, 8, 10]
SKIP_SLIDES = [3, 5, 7, 9, 11]
AUDIO_EXTENSION = "mp3"


def normalize_pid(raw: object) -> str:
    text = str(raw).strip()
    if text.isdigit():
        return str(int(text))
    return text


def ffmpeg_exists() -> bool:
    try:
        subprocess.run(["ffmpeg", "-version"], check=False, capture_output=True)
        return True
    except FileNotFoundError:
        return False


def find_participants(root: Path) -> list[dict]:
    participants = []
    if not root.exists():
        return participants

    for entry in sorted(root.iterdir(), key=lambda p: p.name):
        if not entry.is_dir():
            continue

        m = DIR_PATTERN.match(entry.name)
        if not m:
            continue

        pid = str(int(m.group(1)))
        slide_nums = []
        for wav in entry.glob("*.wav"):
            sm = SLIDE_PATTERN.search(wav.name)
            if sm:
                slide_nums.append(int(sm.group(1)))

        participants.append(
            {
                "id": pid,
                "sessionDir": entry.name,
                "sourceDir": str(entry),
                "availableSlides": sorted(set(slide_nums)),
            }
        )

    participants.sort(key=lambda p: int(p["id"]))
    return participants


def extract_trial_template(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    by_participant: dict[str, list[dict]] = defaultdict(list)

    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        slide = ws.cell(r, 2).value
        trial = ws.cell(r, 3).value
        task = ws.cell(r, 4).value
        item = ws.cell(r, 5).value

        if pid in (None, ""):
            continue
        if slide in (None, "") or trial in (None, "") or task in (None, "") or item in (None, ""):
            continue

        pid_norm = normalize_pid(pid)
        by_participant[pid_norm].append(
            {
                "slide": int(slide),
                "trial": int(trial),
                "task": str(task),
                "item": str(item),
            }
        )

    if not by_participant:
        raise RuntimeError("No trial template rows found in Excel.")

    template_pid, template_rows = max(by_participant.items(), key=lambda kv: len(kv[1]))

    trial_counts: Counter[int] = Counter()
    slide_counts: Counter[int] = Counter()

    template = []
    for i, row in enumerate(template_rows, start=1):
        trial_counts[row["trial"]] += 1
        slide_counts[row["slide"]] += 1
        template.append(
            {
                "index": i,
                "slide": row["slide"],
                "trial": row["trial"],
                "task": row["task"],
                "item": row["item"],
                "itemOrderInTrial": trial_counts[row["trial"]],
                "itemOrderInSlide": slide_counts[row["slide"]],
            }
        )

    print(f"Template source participant: {template_pid} ({len(template)} rows)")
    return template


def reset_audio_root() -> None:
    if AUDIO_DIR.exists() or AUDIO_DIR.is_symlink():
        if AUDIO_DIR.is_symlink() or AUDIO_DIR.is_file():
            AUDIO_DIR.unlink()
        else:
            shutil.rmtree(AUDIO_DIR)
    AUDIO_DIR.mkdir(parents=True, exist_ok=True)


def convert_wav_to_mp3(src_wav: Path, dst_mp3: Path, timeout_sec: int = 20) -> None:
    dst_mp3.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            "ffmpeg",
            "-y",
            "-i",
            str(src_wav),
            "-vn",
            "-codec:a",
            "libmp3lame",
            "-b:a",
            "64k",
            "-ar",
            "22050",
            str(dst_mp3),
        ],
        check=True,
        timeout=timeout_sec,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def build_audio_assets(dataset_name: str, participants: list[dict], source_root: Path) -> tuple[list[dict], int, int]:
    converted = 0
    missing = 0
    manifest_participants = []

    for p in participants:
        session_dir = p["sessionDir"]
        source_dir = source_root / session_dir
        converted_slides = []

        for slide in SCORED_SLIDES:
            stem = f"{session_dir}_slide{slide:02d}"
            src_wav = source_dir / f"{stem}.wav"
            dst_mp3 = AUDIO_DIR / dataset_name / session_dir / f"{stem}.mp3"

            if not src_wav.exists():
                missing += 1
                continue

            try:
                convert_wav_to_mp3(src_wav, dst_mp3)
                converted += 1
                converted_slides.append(slide)
            except subprocess.TimeoutExpired:
                print(f"[timeout] {dataset_name}/{session_dir}/slide{slide:02d}")
                missing += 1
            except subprocess.CalledProcessError:
                print(f"[ffmpeg-error] {dataset_name}/{session_dir}/slide{slide:02d}")
                missing += 1

        manifest_participants.append(
            {
                "id": p["id"],
                "sessionDir": session_dir,
                "availableSlides": converted_slides,
            }
        )

    return manifest_participants, converted, missing


def main() -> None:
    if not ffmpeg_exists():
        raise RuntimeError("ffmpeg not found. Install ffmpeg and rerun.")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    immediate_source = find_participants(IMMEDIATE_SRC)
    delayed_source = find_participants(DELAYED_SRC)
    template = extract_trial_template(SCORES_XLSX)

    reset_audio_root()

    immediate_manifest, immediate_converted, immediate_missing = build_audio_assets(
        dataset_name="immediate",
        participants=immediate_source,
        source_root=IMMEDIATE_SRC,
    )
    delayed_manifest, delayed_converted, delayed_missing = build_audio_assets(
        dataset_name="delayed",
        participants=delayed_source,
        source_root=DELAYED_SRC,
    )

    manifest = {
        "version": "1.1.0",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "scoredSlides": SCORED_SLIDES,
        "skipSlides": SKIP_SLIDES,
        "audioExtension": AUDIO_EXTENSION,
        "scoring": {
            "accuracyValues": [0, 1],
            "partialCredit": False,
            "rule": "Completely correct response = 1, otherwise 0",
        },
        "trialTemplate": template,
        "datasets": [
            {
                "id": "immediate",
                "label": "Immediate",
                "timing": "immediate",
                "audioRoot": "data/audio/immediate",
                "audioExtension": AUDIO_EXTENSION,
                "participants": immediate_manifest,
            },
            {
                "id": "delayed",
                "label": "Delayed",
                "timing": "delayed",
                "audioRoot": "data/audio/delayed",
                "audioExtension": AUDIO_EXTENSION,
                "participants": delayed_manifest,
            },
        ],
    }

    with MANIFEST_PATH.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    total = len({p["id"] for p in immediate_manifest + delayed_manifest})
    print(f"Manifest written: {MANIFEST_PATH}")
    print(f"Immediate: {len(immediate_manifest)} participants, {immediate_converted} mp3 files")
    print(f"Delayed:   {len(delayed_manifest)} participants, {delayed_converted} mp3 files")
    print(f"Total unique participants: {total}")

    total_missing = immediate_missing + delayed_missing
    if total_missing:
        print(f"WARNING: Missing source wav files for {total_missing} scored slides")


if __name__ == "__main__":
    main()
